from nicegui import ui, app
from pathlib import Path
from ytautocontrol.utils.crud import sql
from nicegui.events import ClickEventArguments, GenericEventArguments, UploadEventArguments
from ytautocontrol.utils.socket_handler import SocketHandler, RunScriptFields
import time
from concurrent.futures.thread import ThreadPoolExecutor
from threading import RLock
from ytautocontrol.utils.crud import RunnerScripts
import openpyxl


# ====================== 全局变量 ======================

scripts = {}
selected_script = []
info_dialog: ui.dialog
device_status = []
lock = RLock()
pool = ThreadPoolExecutor()
script_status = {}


# ====================== 全局组件 =======================

card: ui.card


def get_devices_status():
    """获取所有的设备执行状态"""
    global device_status
    device_status = sql.get_devices_running_status()


@ui.refreshable
def info_dialog_component():
    global info_dialog, selected_script
    with ui.dialog().props("full-width") as info_dialog, ui.card().classes("w-full"):
        info_dialog.classes("w-full")
        cols = [
            {"name": "id", "field": "id", "label": "ID", "align": "center"},
            {"name": "device", "field": "device", "label": "设备IP", "align": "center"},
            {"name": "account", "field": "account", "label": "youtube账户", "align": "center"},
            {"name": "word", "field": "word", "label": "搜索关键词", "align": "center"},
            {"name": "author", "field": "author", "label": "视频作者", "align": "center"},
            {"name": "types", "field": "types", "label": "筛选类型", "align": "center"},
            {"name": "freq", "field": "freq", "label": "轮检次数", "align": "center"},
            {"name": "status", "field": "status", "label": "执行状态", "align": "center"},
        ]
        with ui.card_section().classes("flex justify-end w-full"):
            ui.button("关闭", on_click=info_dialog.close, color='red')
        with ui.card_section().classes("w-full"):
            table = ui.table(columns=cols, rows=selected_script).classes("w-full")
            table.add_slot("body-cell-status", """
            <q-td :props="props">
                <q-badge :color="props.value.color">{{ props.value.label }}</q-badge>
            </q-td>
            """)


# def show_info(data):
def show_info(e: GenericEventArguments):
    """显示脚本详细信息按钮回调"""
    global selected_script, device_status
    script_name = e.args["row"]["script_name"]
    selected_script = [{"id": x[0], "device": x[1], "account": x[4], "word": x[7], "author": x[8], "types": x[10] if x[10] else "", "freq": x[11], "status": {}} for x in scripts[script_name]]
    for s in selected_script:
        # sql.cursor.execute(f"select status from running_status where device='{s['device']}';")
        sql.execute(f"select status from running_status where device='{s['device']}';")
        the_status = sql.cursor.fetchone()
        status = {"color": "blue-grey", "label": "未知"}
        if the_status:
            match the_status[0]:
                case 0:
                    status = {"color": "blue", "label": "执行中"}
                case 1:
                    status = {"color": "green", "label": "成功"}
                case 2:
                    status = {"color": "red", "label": "失败"}
                case _:
                    status = {"color": "blue-grey", "label": "未知"}
        s["status"] = status
    info_dialog_component.refresh()
    info_dialog.open()


async def remove_scripts(e: GenericEventArguments):
    global scripts
    script_name = e.args["row"]["script_name"]
    e.sender.props(add="loading")
    result = sql.delete_script_by_name(script_name)
    if result is True:
        ui.notify("删除脚本成功", type='positive')
        scripts = sql.get_scripts_and_group()
        script_cards.refresh()
        main_card.refresh()
    else:
        ui.notify("删除脚本失败", type="negative")
    e.sender.props(remove="loading")


def run_script(e: GenericEventArguments):
    global scripts
    script_name = e.args["row"]["script_name"]
    sql.execute(f"delete from running_status where script_name='{script_name}';")
    for script in scripts[script_name]:
        run_fields: RunScriptFields = {"account": script[4], "password": script[5], "email": script[6], "word": script[7], "author": script[8], "filter_type": script[10], "freq": int(script[11]), "addr": script[1], "script_name": script_name}
        try:
            with SocketHandler(script[1]) as client:
                client.run_script(run_fields)
        except:
            return False
    while True:
        results = sql.execute(f"select device from running_status where script_name='{script_name}';", fetch="all")
        status_devices = [x[0] for x in results if len(x) > 0]
        run_devices = [x[1] for x in scripts[script_name]]
        if len(list(set(run_devices).difference(set(status_devices)))) == 0:
            break
        time.sleep(1)
    e.sender.props(remove="loading")
    script_cards.refresh()
    # main_card.refresh()
    return True


def checker(e: GenericEventArguments):
    global selected_script
    script_name = e.args["row"]["script_name"]
    lock.acquire()
    result = run_script(e)
    lock.release()
    if result is False:
        return False
    while True:
        all_status = sql.get_devices_running_status(script_name)
        if all([x[2] != 0 for x in all_status]):
            lock.acquire()
            script_cards.refresh()
            lock.release()
            return True
        else:
            time.sleep(5)


async def start_script(e: GenericEventArguments):
    e.sender.props(add="loading")
    future = pool.submit(checker, e)


@ui.refreshable
def script_cards():
    global device_status, script_status
    cols = [
        {"name": "script_name", "field": "script_name", "label": "脚本名称", "align": "center"},
        {"name": "status", "field": "status", "label": "执行状态", "align": "center"},
        {"name": "events", "field": "events", "label": "脚本操作", "align": "end"},
    ]
    rows = []
    script_names = [x for x in scripts.keys()]
    for name in script_names:
        all_status = sql.get_devices_running_status(name)
        if all([x[2] != 0 for x in all_status]):
            row = {"script_name": name, "status": "已停止"}
            # script_status[name] = 1
        else:
            row = {"script_name": name, "status": "正在执行"}
            # script_status[name] = 0
        rows.append(row)
    table = ui.table(columns=cols, rows=rows, row_key="name").classes("w-full").props("table-header-class='bg-indigo-100 font-semibold text-lg'")
    table.add_slot("body-cell-events", """
    <q-td :props="props">
        <div class="flex gap-4 justify-end">
            <q-btn @click="$parent.$emit('desc', props)" label="详情" color="primary" flat/>
            <q-btn @click="$parent.$emit('run', props)" icon="not_started" color="primary" round/>
            <q-btn @click="$parent.$emit('remove', props)" icon="delete" color="red" round/>
        </div>
    </q-td>
    """)
    table.on("desc", show_info)
    table.on("run", start_script)
    table.on("remove", remove_scripts)

@ui.refreshable
def main_card():
    with ui.card().classes("w-full"):
        with ui.card_section():
            ui.label("脚本管理").classes("text-lg font-semibold")
            ui.markdown("> 点击脚本卡片详情可以展示当前的脚本详细信息")
            ui.markdown("> 脚本以设备为参照，仅可执行一个脚本，不保证多脚本执行的数据显示正确性，如需多设备执行，请添加对应需求的脚本")
        with ui.card_section().classes("w-full flex gap-4"):
            # model_file = Path(__file__).resolve()
            with ui.button("下载模板", on_click=lambda: ui.download("statics/scripts.xlsx"), icon="download").classes("flex-none").props("outline rounded"):
                ui.tooltip("下载脚本上传模板")
            ui.upload(label="上传xlsx脚本文件", on_upload=script_upload_callback).classes("w-full flex-1")
        with ui.card_section().classes("w-full"):
            ui.separator()
        with ui.card_section().classes("w-full"):
            with ui.column().classes("flex space-y-2 w-full"):
                script_cards()


async def script_upload_callback(e: UploadEventArguments):
    global scripts
    try:
        exist_names = sql.get_scripts_names()
        workbook: openpyxl.Workbook = openpyxl.load_workbook(e.content)
        sheetnames = workbook.sheetnames
        if not len(sheetnames) == len(set(sheetnames)):
            ui.notify(f"脚本名称重复，请检查sheet名称", type="negative")
            return
        # 检测重复的ip和登录账号
        exist_ip_sheets = []
        exist_account_sheets = []
        for sheetname in sheetnames:
            current_sheet_ips = []
            current_sheet_accounts = []
            ws = workbook[sheetname]
            if sheetname in exist_names:
                ui.notify(f"已存在同名脚本：{sheetname}", type="negative")
                return
            for row in ws.iter_rows(values_only=True, min_row=2):
                if row[0] in current_sheet_ips:
                    exist_ip_sheets.append(row[0])
                    break
                else:
                    current_sheet_ips.append(row[0])
                if row[1] in current_sheet_accounts:
                    exist_account_sheets.append(row[1])
                else:
                    current_sheet_accounts.append(row[1])
        if exist_ip_sheets:
            ui.notification(f"同脚本包含重复的IP，重复 IP ：{', '.join(exist_ip_sheets)}", timeout=None, type="negative", close_button=True)
        if exist_account_sheets:
            ui.notification(f"同脚本包含重复的账号，重复账号 ：{', '.join(exist_account_sheets)}", timeout=None, type="negative", close_button=True)
        if any([exist_account_sheets, exist_ip_sheets]):
            return
        successed = []
        faield = []
        for sheetname in sheetnames:
            ws = workbook[sheetname]
            if ws is None:
                ui.notify(f"没有找到表格数据，请确认表格内容", type="negative")
                return
            upload_scripts = []
            for row in ws.iter_rows(values_only=True, min_row=2):
                ip = row[0]
                account = row[1]
                password = row[2]
                email = row[3]
                word = row[4]
                author = row[5]
                types = row[6]
                freq = row[7]
                if not all([ip, account, password, author, freq]):
                    continue
                data: RunnerScripts = {"name": sheetname, "device": ip, "account": account, "password": password, "email": email, "word": word, "author": author, "types": types, "freq": int(freq), "device_account": "", "device_pwd": ""}    # type: ignore
                upload_scripts.append(data)
            result = sql.insert_runner_scripts(upload_scripts)
            if result is True:
                successed.append(sheetname)
            else:
                faield.append(sheetname)
        if faield:
            if successed:
                ui.notify(f"已上传脚本：{', '.join(successed)}", type='positive')
                ui.notify(f"未上传脚本：{', '.join(faield)}", type="negative")
            else:
                ui.notify(f"样本上传失败", type="negative")
        else:
            ui.notify(f"脚本已全部上传", type="positive")
        scripts = sql.get_scripts_and_group()
        script_cards.refresh()
        main_card.refresh()
    except Exception as error:
        ui.notify(f"上传失败: {error}", type="negative")


@ui.page("/runner")
def runner_page():
    global scripts
    with ui.row().classes("fixed top-0 left-0 right-0 bg-indigo-600 w-full h-14 z-10"):
        with ui.row().classes("flex w-full h-full items-center justify-center mx-4"):
            ui.button("YTYOUNB", on_click=lambda: ui.navigate.to("/")).props("flat").classes("flex-1 font-bold text-white text-lg bg-indigo-700 hover:bg-indigo-600/80")
            with ui.row().classes("flex-none"):
                ui.button(icon="menu").props("flat").classes("text-white")
                with ui.menu():
                    # __menu_component()
                    ui.menu_item("示例菜单功能", on_click=lambda: ui.notify("这是留给后续使用的菜单接口"))
    with ui.column().classes("my-20 fixed left-0 ml-4 w-1/4 h-2/4"):
        with ui.card().classes("w-1/2 h-full"):
            ui.label("页面导航").classes("text-lg font-semibold")
            ui.separator()
            # __base_sidebar()
            ui.tree([
                {"id": "/", "label": "创建执行"},
                {"id": "/accounts", "label": "账号管理"},
                {"id": "/devices", "label": "设备管理"},
                {"id": "/runner", "label": "执行管理"},
            ], label_key="label", on_select=lambda e: ui.navigate.to(e.value)).classes("w-full flex flex-col space-y-2")
    with ui.row().classes("w-full h-full"):
        with ui.column().classes("mt-20 mx-80 h-full w-full"):
            get_devices_status()
            info_dialog_component()
            scripts = sql.get_scripts_and_group()
            main_card()


